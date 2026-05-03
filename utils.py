from typing import Callable, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

SEP = "|"
NA = "."

# optional: enforce specific column types
dtype_map = {
    "VEP_Gene_Symbol": "string",
    "Hugo_Symbol (Rahul - not really Hugo)": "string",
}


def load_hugo(file: str) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    """
    Load HUGO gene symbol mapping from HGNC complete set file
    """

    df_hugo = pd.read_csv(
        file,
        sep="\t",
        header=0,
        keep_default_na=False,
    )

    gene_lookup_map = {}
    previous_gene_lookup_map = {}
    alias_gene_lookup_map = {}

    for _, row in df_hugo.iterrows():
        hugo_id = row["HGNC ID"]
        gene_symbol = row["Approved symbol"]

        data = {
            "hgnc_id": hugo_id,
            "approved_symbol": gene_symbol,
        }

        refseqs = (
            [x.strip() for x in row["RefSeq IDs"].split(",")]
            if row["RefSeq IDs"]
            else []
        )
        ensembl_ids = (
            [x.strip() for x in row["Ensembl gene ID"].split(",")]
            if row["Ensembl gene ID"]
            else []
        )

        gene_lookup_map[hugo_id.lower()] = data
        gene_lookup_map[gene_symbol.lower()] = data
        for refseq in refseqs:
            gene_lookup_map[refseq.lower()] = data
        for ensembl_id in ensembl_ids:
            gene_lookup_map[ensembl_id.lower()] = data

        previous_symbols = (
            [x.strip() for x in row["Previous symbols"].split(",")]
            if row["Previous symbols"]
            else []
        )
        for prev in previous_symbols:
            previous_gene_lookup_map[prev.lower()] = data

        alias_symbols = (
            [x.strip() for x in row["Alias symbols"].split(",")]
            if row["Alias symbols"]
            else []
        )
        for alias in alias_symbols:
            alias_gene_lookup_map[alias.lower()] = data

    return {
        "current": gene_lookup_map,
        "previous": previous_gene_lookup_map,
        "alias": alias_gene_lookup_map,
    }


def get_is_hugo_gene(
    gene_id: str,
    symbol: str,
    hugo_info: dict[str, dict[str, dict[str, str]]],
) -> dict[str, Union[str, int]]:
    gene_id_lower = gene_id.lower()
    symbol_lower = symbol.lower()

    ret = {"gene_id": gene_id, "symbol": symbol, "hgnc_id": NA, "is_hugo_gene": 0}

    gene_info = None

    if gene_id_lower in hugo_info["current"]:
        gene_info = hugo_info["current"][gene_id_lower]
    elif gene_id_lower in hugo_info["previous"]:
        gene_info = hugo_info["previous"][gene_id_lower]
    elif gene_id_lower in hugo_info["alias"]:
        gene_info = hugo_info["alias"][gene_id_lower]
    elif symbol_lower in hugo_info["current"]:
        gene_info = hugo_info["current"][symbol_lower]
    elif symbol_lower in hugo_info["previous"]:
        gene_info = hugo_info["previous"][symbol_lower]
    elif symbol_lower in hugo_info["alias"]:
        gene_info = hugo_info["alias"][symbol_lower]
    else:
        gene_info = None

    if gene_info is not None:
        ret["hgnc_id"] = gene_info["hgnc_id"]
        ret["symbol"] = gene_info["approved_symbol"]
        ret["is_hugo_gene"] = 1

    return ret


def load_transcripts(file: str, transcript_map) -> dict[str, dict]:
    """
    Load transcript mapping from GENCODE basic transcript file
    """

    df = pd.read_csv(file, sep="\t", header=0, keep_default_na=False)

    for _, row in df.iterrows():
        gene_id = row["gene_id"]
        gene_symbol = row["gene_symbol"]
        transcript_id = row["transcript_id"]
        appris = row["appris"]
        # strip version from ccds_id if exists
        ccds = row["ccds"].split(".")[0]
        is_canonical = row["is_canonical"]

        if transcript_id not in transcript_map:
            transcript_map[transcript_id] = {
                "gene_id": gene_id,
                "gene_symbol": gene_symbol,
                "appris": appris if appris != "" else NA,
                "ccds": ccds if ccds != "" else NA,
                "is_canonical": is_canonical,
            }

        # update fields to get rid of NAs

        if ccds != NA and ccds != "":
            transcript_map[transcript_id]["ccds"] = ccds

        if is_canonical == 1:
            transcript_map[transcript_id]["is_canonical"] = 1

        if appris != NA and appris != "":
            transcript_map[transcript_id]["appris"] = appris

    return transcript_map


# def load_v19_transcripts(file: str) -> dict[str, dict]:
#     """
#     Load transcript mapping from GENCODE basic transcript file
#     """

#     df = pd.read_csv(file, sep="\t", header=0, keep_default_na=False)

#     transcript_map = {}

#     for _, row in df.iterrows():
#         gene_id = row["gene_id"]
#         gene_symbol = row["gene_symbol"]
#         hugo_gene_symbol = row["hugo_gene_symbol"]
#         transcript_id = row["transcript_id"]
#         appris = row["appris"]
#         # strip version from ccds_id if exists
#         ccds = row["ccds"].split(".")[0]
#         is_canonical = row["is_canonical"]

#         transcript_map[transcript_id] = {
#             "gene_id": gene_id,
#             "gene_symbol": gene_symbol,
#             "hugo_gene_symbol": hugo_gene_symbol,
#             "appris": appris if appris != "" else NA,
#             "ccds": ccds if ccds != "" else NA,
#             "is_canonical": is_canonical,
#         }

#     return transcript_map


def load_ccds_lengths(file: str) -> dict[str, dict]:
    """
    Load CCDS lengths from file from NCBI CCDS project,
    using the protein lengths so we don't need to calculate
    from genomic coordinates and exon structures,
    which can be error prone using divide by 3
    """
    df = pd.read_csv(file, sep="\t", header=0, keep_default_na=False)

    print(f"Loaded {len(df)} CCDS entries from {file}")
    df = df[df["aa_length"] != -1]
    print(f"{len(df)} CCDS entries with valid aa_length")

    ccds_length_map = {}

    for _, row in df.iterrows():
        # strip version from ccds_id if exists
        ccds = row["ccds_id"].split(".")[0]
        aa_length = row["aa_length"]

        ccds_length_map[ccds] = {
            "aa_length": aa_length,
        }

    # print(ccds_length_map)

    return ccds_length_map


def maf_to_excel(df, fout):
    df = df.copy()

    for col, typ in dtype_map.items():
        if col in df.columns:
            df[col] = df[col].astype(typ)

    df.to_excel(fout, index=False)

    arial = Font(name="Arial")
    # define font
    header_font = Font(name="Arial", bold=True)

    # reopen workbook
    wb = load_workbook(fout)
    ws = wb.active

    for col in ws.columns:
        for cell in col:
            cell.font = arial

    # apply to header row
    for cell in ws[1]:
        cell.font = header_font

    wb.save(fout)


def chr_sizes(assembly: str = "hg19"):
    df_sizes = pd.read_csv(
        f"/ifs/archive/cancer/Lab_RDF/scratch_Lab_RDF/ngs/references/ucsc/{assembly}_chromosome_sizes.txt",
        sep="\t",
        header=0,
    )

    sizes = [
        {"chr": row["chrom"], "size": row["size"]} for _, row in df_sizes.iterrows()
    ]

    return sizes


def chunk_table(
    fin: str, fout: str, fn: Callable[[pd.DataFrame], None], chunksize=200000
):
    first = True
    chunk = 1

    for df in pd.read_csv(
        fin,
        sep="\t",
        header=0,
        keep_default_na=False,
        chunksize=chunksize,
    ):
        print(f"Processing chunk {chunk}...")
        chunk += 1

        fn(df)

        if fout:
            df.to_csv(
                fout,
                sep="\t",
                header=first,
                mode="w" if first else "a",
                index=False,
            )

        first = False
